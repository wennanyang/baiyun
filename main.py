import win32com.client as win32
import os
import glob
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
from tqdm import tqdm
import traceback
import xlrd3 as xlrd
import re
import tkinter as tk
from tkinter import filedialog
import threading
from tkinter import ttk
def check_dir(dir):
    dir = os.path.join('异常文件汇总', dir)
    if not os.path.exists(dir):
        os.makedirs(dir)
        return dir
    shutil.rmtree(dir)
    os.makedirs(dir)
    return dir
def check_file(file, sheet_name="放线数据汇总"):
    file = os.path.join(os.getcwd(), file)
    if os.path.exists(file):
        os.remove(file)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(file)
def find_match_files_recursion(parent_dir, re_pattern):
    file_set = set()
    file_list = []
    for root, _, files in os.walk(parent_dir):
        for file in files:
            if re.match(re_pattern, file) is not None and file not in file_set:
                file_set.add(file)
                file_list.append(os.path.join(root, file))
    return file_list
def make_fu_result(fu_dir, sheet_name="验收成果汇总", save_name= "验收成果汇总表.xlsx",
                title=["工程编号", "建筑结构", "建设单位", "建设项目名称", "建设位置", 	
                       "建设工程规划许可证号", "相关批文号", "放线案号", "建设规模", 
                       "基底面积(m2)", "住宅户数", "汽车泊位(个)", "总建筑面积(m2)", 
                        "地上面积(m2)", "地下面积(m2)", "地上层数", "地下层数", 
                        "主要功能", "建筑高度(m)", "更新时间", "备注"],
                exception_filename='log.txt',
                exp_doc_name="验收提取异常的doc",
                exp_xls_name="验收提取异常的xls",
                empty_xls_name="验收提取为空的xls",
                exception_check_dir = "异常的验收项目",
                progress_callback=None):
    check_file(save_name, sheet_name=sheet_name)
    copy_filename = "复制的文件名.txt"
    if os.path.exists(exception_filename):
        os.remove(exception_filename)
    if os.path.exists(copy_filename):
        os.remove(copy_filename)
    exp_doc_dir = check_dir(exp_doc_name)
    exp_xls_dir = check_dir(exp_xls_name)
    empty_xls_dir = check_dir(empty_xls_name)
    exception_check_project_set = set()
    
    os.makedirs(exception_check_dir,exist_ok=True)
    project_dir = glob.glob(os.path.join(fu_dir, "*"))
    exception_doc_list = []
    exception_xls_list = []
    empty_xls_list = []
    project_fu_list = []
    wb = load_workbook(save_name)
    ws = wb[sheet_name]
    ws.append(title)
    ## 总的异常数
    exp_count = 0
    tech_check_re = r"^[^~]*技术审查.*\.xls[x]?$"
    achieve_re = r'^[^~]*成果.*\.doc'
    for i , project in tqdm(enumerate(project_dir), total=len(project_dir)):
        tech_check_list = find_match_files_recursion(project, tech_check_re)
        buildings_high = ""
        ## 该循环下的异常数，如果为0，说明顺利完成
        count = 0
        for tech_check in tech_check_list:
            try:
                high = get_buildings_high(tech_check=tech_check)
                if high == None:
                    empty_xls_list.append(tech_check)
                    continue
                else :
                    buildings_high += high
                    break
            except Exception as e:
                ## 写入日志
                with open(exception_filename, 'a', encoding='utf-8') as f:
                    f.write(f"{exp_count}'\t'{tech_check}'\n'{traceback.format_exc()}'\n'")
                    exp_count += 1
                count += 1
                ## 将异常的文件挪出来
                exception_check_project_set.add(project)
                exception_xls_list.append(tech_check)
                continue
        doc_list = find_match_files_recursion(project, achieve_re)
        for doc_path in doc_list:
            
            try:
                result = get_doc_result(doc_path=doc_path)
            except Exception as e:
                ## 写入日志
                with open(exception_filename, 'a', encoding='utf-8') as f:
                    f.write(f"{exp_count}'\t'{doc_path}'\n'{traceback.format_exc()}'\n'")
                    exp_count+=1
                ## 将异常的文件挪出来
                exception_doc_list.append(doc_path)
                exception_check_project_set.add(project)
                count += 1
                continue

            result[18] = buildings_high
            ## 如果异常数为0，那么这个就是已经被提取完成的项目，加入完成项目的列表
            if count == 0:
                project_fu_list.append(result[0])
            ws.append(result)
        ## 如果审查或者成果表都为0，说明出现问题了，加入异常项目列表
        if len(doc_list) == 0 or len(tech_check_list) == 0:
            exception_check_project_set.add(project)
        ## 回调函数，设置进度条
        if progress_callback is not None:
            progress_callback((i + 1) / len(project_dir))
    copy_filename_list = exception_doc_list + exception_xls_list + empty_xls_list
    if len(copy_filename_list) > 0:
        with open(copy_filename, 'w+') as f:
            for item in copy_filename_list:
                f.write(item + "\n")

    for exception_doc in exception_doc_list:   
        copy_name = os.path.basename(os.path.dirname(exception_doc)) + '-' + os.path.basename(exception_doc)
        try:
            shutil.copy(exception_doc, os.path.join(exp_doc_dir, copy_name))
        except Exception as e:
            print(e)
            continue
    for exception_xls in exception_xls_list:   
        copy_name = os.path.basename(os.path.dirname(exception_xls)) + '-' + os.path.basename(exception_xls)
        try:
            shutil.copy(exception_xls, os.path.join(exp_xls_dir, copy_name))
        except Exception as e:
            print(e)
            continue
    for empty_xls in empty_xls_list:
        copy_name = os.path.basename(os.path.dirname(empty_xls)) + '-' + os.path.basename(empty_xls)
        try:
            shutil.copy(empty_xls, os.path.join(empty_xls_dir, copy_name))
        except Exception as e:
            print(e)
    for exception_project in exception_check_project_set:
        try:
            print(exception_check_dir)
            print(exception_check_project_set)
            shutil.copytree(exception_project, os.path.join(exception_check_dir, os.path.basename(exception_project)))
        except Exception as e:
            print(e)   
    wb.save(save_name)
    return project_fu_list

def make_fang_result(fang_dir, exp_dir="放线提取异常的xls",sheet_name="放线数据汇总", save_name= "放线数据汇总.xlsx",
                    title=["工程编号","建筑结构","建设单位","建设项目名称","建设位置",
                       "建设工程规划许可证号","更新时间","备注"],
                    exception_filename='放线异常的文件列表.txt',
                    exception_fang_dir = "异常的放线项目",
                    progress_callback=None):
    check_file(save_name, sheet_name=sheet_name)  
    if os.path.exists(exception_filename):
        os.remove(exception_filename)

    excels = glob.glob(os.path.join(fang_dir, "*\\*放*.xls*"))
    excel_list = [s for s in excels if not os.path.basename(s).startswith('~')]
    wb = load_workbook(save_name)
    ws = wb[sheet_name]
    ws.append(title)
    ws.title= sheet_name
    exp_count = 1
    project_fang_list = []
    os.makedirs(exception_fang_dir,exist_ok=True)  
    for i, excel_path in enumerate(excel_list):
        count = 0
        try:
            result = get_fang_result(excel_path=excel_path)
        except Exception as e:
            with open(exception_filename, 'a', encoding='utf-8') as f:
                f.write(f"{exp_count}'\t'{excel_path}'\n'{traceback.format_exc()}")
                exp_count+=1
            
            copy_name = os.path.basename(os.path.dirname(excel_path)) + '-' + excel_path
            shutil.copy(excel_path, os.path.join(exp_dir, copy_name))
            project_name = os.path.dirname(excel_path)
            shutil.copytree(project_name, os.path.join(exception_fang_dir, os.path.basename(project_name)))
            count += 1
            continue
        ws.append(result)
        if count == 0:
            project_fang_list.append(result[0])
        if progress_callback is not None:
            progress_callback(i / len(excel_list) * 100)

    wb.save(save_name)
    return project_fang_list

def get_fang_result(excel_path):
    result_list = [""] * 8
    wb = xlrd.open_workbook(excel_path)
    if "放线" in wb.sheet_names():
        sheet = wb.sheet_by_name('放线')
        result_list[0] = sheet.cell_value(1, 1)
        result_list[2] = sheet.cell_value(2, 1)
        result_list[3] = sheet.cell_value(4, 1)
        result_list[4] = sheet.cell_value(3, 1)
    return result_list

def get_doc_result(doc_path):
    word = win32.Dispatch("Word.Application")
    word.visible = False
    wb = word.Documents.Open(doc_path)
    doc = word.ActiveDocument
    # word.Quit()
    result_list = [""] * 21
    table1 = doc.Tables(1)
    # 0 工程编号
    result_list[0] = os.path.basename(os.path.dirname(doc_path))
    # 1 建筑结构
    result_list[1] = ""
    # 2 建设单位
    result_list[2] = table1.Cell(Row=1, Column=2).Range.Text
    # 3 建设项目名称
    result_list[3] = table1.Cell(Row=2, Column=2).Range.Text
    # 4 建设位置
    result_list[4] = table1.Cell(Row=3, Column=2).Range.Text
    # 5 建设工程规划许可证号
    result_list[5] = table1.Cell(Row=7, Column=2).Range.Text
    # 6 相关批文号
    result_list[6] = table1.Cell(Row=8, Column=2).Range.Text
    # 7 放线案号
    result_list[7] = table1.Cell(Row=9, Column=2).Range.Text.split('/')[0]
    # 8 建设规模
    result_list[8] = table1.Cell(Row=11, Column=2).Range.Text
    # 9 基底面积
    result_list[9] = table1.Cell(Row=13, Column=4).Range.Text
    # 10 住宅户数
    result_list[10] = table1.Cell(Row=17, Column=4).Range.Text
    # 11 汽车泊位
    result_list[11] = table1.Cell(Row=16, Column=4).Range.Text
    table2 = doc.Tables(2)
    # 12 总建筑面积
    result_list[12] = table2.Cell(Row=2, Column=5).Range.Text
    # 13 地上面积
    result_list[13] = table2.Cell(Row=3, Column=6).Range.Text
    # 14地下面积
    result_list[14] = table2.Cell(Row=4, Column=6).Range.Text
    # 15 地上层数
    result_list[15] = table2.Cell(Row=5, Column=5).Range.Text
    # 16 地下层数
    result_list[16] = table2.Cell(Row=6, Column=5).Range.Text
    # 17 主要功能
    result_list[17] = table2.Cell(Row=10, Column=2).Range.Text + \
                        table2.Cell(Row=11, Column=2).Range.Text + \
                        table2.Cell(Row=12, Column=2).Range.Text
    # 18 建筑高度
    result_list[18] = "" 
    # 19 更新时间
    result_list[19] = ""
    # 20备注
    result_list[20] = ""

    result = [s.replace('\r', '').replace('\x07', '') for s in result_list]
    doc.Close(SaveChanges=False)
    word.Quit()
    return result
def get_buildings_high(tech_check):
    building_high = None
    wb = xlrd.open_workbook(tech_check)
    if "验收" in wb.sheet_names():
        ws = wb.sheet_by_name("验收")
        if ws.cell_value(27, 1) != None and "建筑高度" in str(ws.cell_value(26, 1)):
            building_high = str(ws.cell_value(26, 5))
        elif ws.cell_value(27, 1) != None and "建筑高度" in str(ws.cell_value(26, 2)):
            building_high = str(ws.cell_value(26, 6))
        elif ws.cell_value(27, 1) != None and "建筑高度" in str(ws.cell_value(27, 1)):
            building_high = str(ws.cell_value(27, 5))
        elif ws.cell_value(27, 1) != None and "建筑高度" in str(ws.cell_value(27, 2)):
            building_high = str(ws.cell_value(27, 6))
    return building_high

def validate_project(path_xls, project_fang_list, project_fu_list, filtered_file_name):
    project_list_total = []
    project_list = project_fang_list + project_fu_list
    wb = xlrd.open_workbook(path_xls)
    sheet = wb.sheet_by_index(0)
    column_index = 2
    for row in range(1, sheet.nrows):
        project_list_total.append(sheet.cell_value(row, column_index))
    filtered = [item for item in project_list_total if item not in project_list]
    if (len(filtered) > 0) : 
        with open(filtered_file_name, 'w+', encoding='utf-8') as f:
            for item in filtered:
                f.write(item+'\n')

def main(fang_dir=None, fu_dir=None, validate_xls=None, progress_callback=None, args=None):
    fangxls = '放线提取异常的xls'
    fuxls = '验收提取异常的xls'
    fudoc = '验收提取异常的doc'
    fuempty = '验收提取为空的xls'
    project_fang_list = []
    project_fu_list = []
    if fang_dir is not None:
        exp_fang_dir = check_dir(fangxls)
        project_fang_list = make_fang_result(fang_dir=fang_dir, exp_dir=exp_fang_dir,
                                             progress_callback=progress_callback)
    if fu_dir is not None:
        exp_xls_dir = check_dir(fuxls)
        exp_doc_dir = check_dir(fudoc)
        empty_xls_dir = check_dir(fuempty)
        project_fu_list = make_fu_result(fu_dir=fu_dir, exp_xls_dir=exp_xls_dir, 
                                         exp_doc_dir=exp_doc_dir, 
                                         empty_xls_dir=empty_xls_dir,
                                         progress_callback=progress_callback)

    if validate_xls is not None:
        validate_project(path_xls=validate_xls, project_fu_list=project_fu_list, 
                        project_fang_list=project_fang_list, 
                        filtered_file_name="放线验收缺失的项目列表.txt")
class RedirectText:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)  # 自动滚动到最后一行

    def flush(self):
        pass  # 这个方法不需要做任何事情，但需要存在
class GUI():
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.fang_var = tk.StringVar()
        self.fu_var = tk.StringVar()
        self.validate_var = tk.StringVar()
        self.root.title("放线验收数据程序")
        self.root.configure(bg="skyblue")
        self.root.minsize(200, 200)  # width, height
        self.root.maxsize(500, 500)
        self.root.geometry("600x300+250+250")
        fang_frame = tk.Frame(self.root, bg="#6FAFE7")
        # 设置第一行放线路径的label, entry, button
        fang_frame.grid(row=0, column=0, padx=10, pady=10, sticky='nsew')
        fang_label = tk.Label(fang_frame, text="放线路径", bg="#6FAFE7")
        fang_label.grid(row=0, column=0)
        self.fang_entry = tk.Entry(fang_frame, bd=3, width=50, textvariable=self.fang_var)
        self.fang_entry.grid(row=0, column=1)
        choose_fang_dir = tk.Button(fang_frame, text="选择文件", 
                                    command=lambda: self.select_directory(self.fang_var))
        choose_fang_dir.grid(row=0, column=2)
        # 设置第二行验收路径的label, entry, button
        fu_frame = tk.Frame(self.root, bg="#6FAFE7")
        fu_frame.grid(row=1, column=0, padx=10, pady=10, sticky='nsew')
        fu_lable = tk.Label(fu_frame, text="验收路径", bg="#6FAFE7")
        fu_lable.grid(row=0, column=0)
        self.fu_entry = tk.Entry(fu_frame, bd=3, width=50,textvariable=self.fu_var)
        self.fu_entry.grid(row=0, column=1)
        choose_fu_dir = tk.Button(fu_frame, text="选择目录", 
                                  command=lambda: self.select_directory(self.fu_var))
        choose_fu_dir.grid(row=0, column=2)
        # 设置第三行的验证excel路径
        validate_frame = tk.Frame(self.root, bg="#6FAFE7")
        validate_frame.grid(row=2, column=0, padx=10, pady=10, sticky='nsew')
        validate_lable = tk.Label(validate_frame, text="清单路径", bg="#6FAFE7")
        validate_lable.grid(row=0, column=0)
        self.validate_entry = tk.Entry(validate_frame, bd=3, width=50, textvariable=self.validate_var)
        self.validate_entry.grid(row=0, column=1)
        choose_validate = tk.Button(validate_frame, text="选择文件", 
                                    command=lambda: self.select_file(self.validate_var))
        choose_validate.grid(row=0, column=2)
        # 设置第四行验收路径的label, entry, button
        self.turn_on = tk.Button(self.root, text="开始执行", command=self.excute)
        self.turn_on.grid(row=3)

        ## 进度条
        self.progress_var = tk.IntVar()
        self.progressbar = ttk.Progressbar(self.root, orient="horizontal", length=300, mode="determinate", 
                                              variable=self.progress_var)
        self.progressbar.grid(row=4)
    def mainloop(self):
        self.root.mainloop()
    def select_directory(self, stringvar):
        directory = filedialog.askdirectory()
        if directory:
            stringvar.set(directory)
    def select_file(self, stringvar):
        # 打开文件选择对话框并获取文件路径
        file_path = filedialog.askopenfilename(
            title="选择一个xls文件",  # 对话框标题
            filetypes=[("excel文件", "*.xls"), ("excel文件", "*.xlsx") ]  # 过滤文件类型
        )
        if file_path:
            stringvar.set(file_path)
    def excute(self):
        print(self.fang_entry.get())
        print(self.fu_entry.get())
        print(self.validate_entry.get())
        threading.Thread(target=self.long_running_task,args=(), daemon=True).start()
    def update_progress(self, value):
        self.progress_var.set(value)
        self.root.update_idletasks() 
    def long_running_task(self):
        main(fang_dir=self.fang_entry.get(), 
             fu_dir=self.fu_entry.get(), 
             validate_xls=self.validate_entry.get(),
             progress_callback=self.update_progress)
if __name__ == '__main__':
    # gui = GUI()
    # gui.mainloop()
    # file_list =  find_match_files_recursion(r'F:\2023复23B059',r"^[^~]*技术审查.*\.xls[x]?$")
    # print(file_list)
    make_fu_result(fu_dir=r"F:\验收")

