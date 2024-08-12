import win32com.client as win32
import shutil
from openpyxl import load_workbook
import traceback
import xlrd3 as xlrd
import re
from pathlib import Path
from utils import check_dir, check_file, find_match_files_recursion, find_match_txt_recursion,ignore_hidden_files
EXP_BASE_DIR = Path('异常文件汇总')
EXP_BASE_DIR.mkdir(parents=True, exist_ok=True)

def make_fu_result(fu_dir : Path, sheet_name="验收成果汇总", save_name=Path("验收成果汇总表.xlsx"),
                exception_check_dir = Path("异常的验收项目"),
                progress_callback=None):
    title=["工程编号", "建筑结构", "建设单位", "建设项目名称", "建设位置", 	
        "建设工程规划许可证号", "相关批文号", "放线案号", "建设规模", 
        "基底面积(m2)", "住宅户数", "汽车泊位(个)", "总建筑面积(m2)", 
        "地上面积(m2)", "地下面积(m2)", "地上层数", "地下层数", 
        "主要功能", "建筑高度(m)", "更新时间", "备注"]
    exception_check_dir.mkdir(parents=True, exist_ok=True)
    copy_name = "复制的文件名.txt"
    log_name='log.txt'
    exp_doc_name="验收提取异常的doc"
    exp_xls_name="验收提取异常的xls"
    empty_xls_name="验收提取为空的xls"
    check_file(save_name, sheet_name=sheet_name)
    log_filename = EXP_BASE_DIR.joinpath(log_name)
    if log_filename.exists():
        log_filename.unlink()
    copy_filename = EXP_BASE_DIR.joinpath(copy_name)
    if copy_filename.exists():
        copy_filename.unlink()
    exp_doc_dir = check_dir(Path(exp_doc_name))
    exp_xls_dir = check_dir(Path(exp_xls_name))
    empty_xls_dir = check_dir(Path(empty_xls_name))
    exception_check_project_set = set()
    project_dir = list(fu_dir.glob("*"))
    exception_doc_list = []
    exception_xls_list = []
    empty_xls_list = []
    wb = load_workbook(save_name)
    ws = wb[sheet_name]
    ws.append(title)
    ## 总的异常数
    exp_count = 0
    tech_check_re = r"^[^~]*技术审查.*\.xls[x]?$"
    achieve_re = r'^[^~]*成果.*\.doc'
    for i, project in enumerate(project_dir):
        project = project.resolve()
        tech_check_list = find_match_files_recursion(Path(project), tech_check_re,
                                                     suffix="*.xls*")
        buildings_high = ""
        for tech_check in tech_check_list:
            try:
                high = get_buildings_high(tech_check=tech_check)
                if high == None:
                    empty_xls_list.append(tech_check)
                    continue
                else :
                    buildings_high += high
                    break
            except Exception as e:
                ## 写入日志
                with open(log_filename, 'a', encoding='utf-8') as f:
                    f.write(f"{exp_count}'\t'{tech_check}'\n'{traceback.format_exc()}'\n'")
                    exp_count += 1
                ## 将异常的文件挪出来
                exception_check_project_set.add(project)
                exception_xls_list.append(tech_check)
                continue
        doc_list = find_match_files_recursion(project, achieve_re)
        for doc_path in doc_list:
            try:
                result = get_doc_result(doc_path=doc_path)
            except Exception as e:
                ## 写入日志
                with open(log_filename, 'a', encoding='utf-8') as f:
                    f.write(f"{exp_count}'\t'{doc_path}'\n'{traceback.format_exc()}'\n'")
                    exp_count+=1
                ## 将异常的文件挪出来
                exception_doc_list.append(doc_path)
                exception_check_project_set.add(project)
                continue

            result[18] = buildings_high
            ws.append(result)
        ## 如果审查或者成果表都为0，说明出现问题了，加入异常项目列表
        if len(doc_list) == 0 and len(tech_check_list) == 0:
            exception_check_project_set.add(project)
        ## 回调函数，设置进度条
        if progress_callback is not None:
            progress_callback((i + 1) / len(project_dir) * 100, 
                              description="正在提取验收属性")
    copy_filename_list = exception_doc_list + exception_xls_list + empty_xls_list
    if len(copy_filename_list) > 0:
        with open(copy_filename, 'w+') as f:
            for item in copy_filename_list:
                f.write(str(item) + "\n")

    for exception_doc in exception_doc_list:   
        copy_name = exception_doc.parent.name + '-' + exception_doc.name
        try:
            shutil.copy(exception_doc, exp_doc_dir.joinpath(copy_name))
        except Exception as e:
            print(e)
            continue
    for exception_xls in exception_xls_list:   
        copy_name = exception_xls.parent.name + '-' + str(exception_xls.name)
        try:
            shutil.copy(exception_xls, exp_xls_dir.joinpath(copy_name))
        except Exception as e:
            print(e)
            continue
    for empty_xls in empty_xls_list:
        copy_name = empty_xls.parent.name + '-' + empty_xls.name
        try:
            shutil.copy(empty_xls, empty_xls_dir.joinpath(copy_name))
        except Exception as e:
            print(e)
            continue
    for exception_project in exception_check_project_set:
        try:
            shutil.copytree(exception_project, 
                            exception_check_dir.joinpath(exception_project.name),
                            ignore=ignore_hidden_files,
                            dirs_exist_ok=True)
        except Exception as e:
            print(e) 
            continue  
    wb.save(save_name)

def make_fang_result(fang_dir : Path, exp_dir_name="放线提取异常的txt",sheet_name="放线数据汇总",
                     save_name=Path("放线数据汇总.xlsx"),progress_callback=None):
    title=["工程编号","建筑结构","建设单位","建设项目名称","建设位置",
        "放线依据", "用地许可证号", "更新时间","备注"]
    exception_name='放线异常的文件列表.txt'
    empty_name = "放线没有txt的项目.txt"
    exception_fang_dir = Path("异常的放线项目")
    check_file(save_name, sheet_name=sheet_name) 
    exception_filename = EXP_BASE_DIR.joinpath(exception_name) 
    empty_filename = EXP_BASE_DIR.joinpath(empty_name)
    empty_txt_project = []
    if exception_filename.exists():
        exception_filename.unlink()
    exception_fang_dir.mkdir(parents=True, exist_ok=True)
    exp_dir = check_dir(exp_dir_name) 
    project_list = list(fang_dir.glob("*"))
    wb = load_workbook(save_name)
    ws = wb[sheet_name]
    ws.append(title)
    ws.title= sheet_name
    exp_count = 1
    for i, project in enumerate(project_list):
        project = project.resolve()
        txt_path = find_match_txt_recursion(project, r'.*\.(?i:txt)$')
        if txt_path is None:
            empty_txt_project.append(project)
            continue
        try:
            result = get_fang_result_from_txt(txt_path)
            if result is None:
                shutil.copy(txt_path, exp_dir.joinpath((txt_path.name)))
                continue
        except Exception as e:
            with open(exception_filename, 'a', encoding='utf-8') as f:
                f.write(f"{exp_count}'\t'{txt_path}'\n'{traceback.format_exc()}")
                exp_count+=1
            
            copy_name = project.name + '-' + txt_path.name
            shutil.copy(txt_path, exp_dir.joinpath(copy_name))
            shutil.copytree(project, 
                            exception_fang_dir.joinpath(project.name),
                            ignore=ignore_hidden_files,
                            dirs_exist_ok=True)
            continue
        ws.append(result)
        if progress_callback is not None:
            progress_callback(i / len(project_list) * 100, description="正在提取放线属性")

    wb.save(save_name)
    with open(empty_filename, 'w+', encoding='utf-8') as f:
        for empty in empty_txt_project:
            f.write(str(empty) + '\n')

def suply_make_fang(dir : Path):
    txt_path_list = dir.glob("*")
    save_name = "放线数据汇总.xlsx"
    wb = load_workbook(save_name)
    ws = wb["放线数据汇总"]
    for txt_path in txt_path_list:
        try:
            result = get_fang_result_from_txt(txt_path.resolve())
            if result is None:
                shutil.copy(txt_path, Path('异常文件汇总\放线提取异常的txt').joinpath(txt_path.name))
                continue
        except Exception as e:
            print(e)
        ws.append(result)
    wb.save(save_name)
    
def get_fang_result(excel_path):
    result_list = [""] * 8
    wb = xlrd.open_workbook(excel_path)
    if "放线" in wb.sheet_names():
        sheet = wb.sheet_by_name('放线')
        result_list[0] = sheet.cell_value(1, 1)
        result_list[2] = sheet.cell_value(2, 1)
        result_list[3] = sheet.cell_value(4, 1)
        result_list[4] = sheet.cell_value(3, 1)
    return result_list
def get_fang_result_from_txt(txt_path):
    result = [""] * 9
    with open(txt_path, 'r', encoding='GBK') as f:
        lines = f.readlines()
    pattern = r'\d{4}[放F]\d{2}[A-Z]\d{3}'
    for i in range(22, len(lines)):
        project_name = lines[i].split(":")[-1].strip()
        match = re.match(pattern=pattern, string=project_name)
        if match is not None:
            result[0] = project_name
            break
        if i == len(lines) - 1:
            return None
    
    result[1] = ""
    result[2] = lines[5].split(":")[-1].strip()
    result[3] = lines[2].split(":")[-1].strip()
    result[4] = lines[4].split(":")[-1].strip()
    result[5] = lines[0].split(":")[-1].strip()
    result[6] = lines[1].split(":")[-1].strip()
    return result

            
def get_doc_result(doc_path : Path):
    word = win32.Dispatch("Word.Application")
    word.visible = False
    wb = word.Documents.Open(str(doc_path))
    doc = word.ActiveDocument
    # word.Quit()
    result_list = [""] * 21
    table1 = doc.Tables(1)
    # 0 工程编号
    pattern = r'\d{4}复\d{2}[A-Z]\d{3}'
    if re.match(pattern, doc_path.parent.name):
        result_list[0] = doc_path.parent.name
    # elif re.match(pattern, table1.Cell(Row=9, Column=2).Range.Text) :
    #     r
    else: 
        match = re.search(pattern, doc_path)
        if match is not None:
            result_list[0] = match.group(0)
        else :
            result_list[0] = doc_path
    # 1 建筑结构
    result_list[1] = ""
    # 2 建设单位
    result_list[2] = table1.Cell(Row=1, Column=2).Range.Text
    # 3 建设项目名称
    result_list[3] = table1.Cell(Row=2, Column=2).Range.Text
    # 4 建设位置
    result_list[4] = table1.Cell(Row=3, Column=2).Range.Text
    # 5 建设工程规划许可证号
    result_list[5] = table1.Cell(Row=7, Column=2).Range.Text
    # 6 相关批文号
    result_list[6] = table1.Cell(Row=8, Column=2).Range.Text
    # 7 放线案号
    result_list[7] = table1.Cell(Row=9, Column=2).Range.Text.split('/')[0]
    # 8 建设规模
    result_list[8] = table1.Cell(Row=11, Column=2).Range.Text
    # 9 基底面积
    result_list[9] = table1.Cell(Row=13, Column=4).Range.Text
    # 10 住宅户数
    result_list[10] = table1.Cell(Row=17, Column=4).Range.Text
    # 11 汽车泊位
    result_list[11] = table1.Cell(Row=16, Column=4).Range.Text
    table2 = doc.Tables(2)
    # 12 总建筑面积
    result_list[12] = table2.Cell(Row=2, Column=5).Range.Text
    # 13 地上面积
    result_list[13] = table2.Cell(Row=3, Column=6).Range.Text
    # 14地下面积
    result_list[14] = table2.Cell(Row=4, Column=6).Range.Text
    # 15 地上层数
    result_list[15] = table2.Cell(Row=5, Column=5).Range.Text
    # 16 地下层数
    result_list[16] = table2.Cell(Row=6, Column=5).Range.Text
    # 17 主要功能
    result_list[17] = table2.Cell(Row=10, Column=2).Range.Text + \
                        table2.Cell(Row=11, Column=2).Range.Text + \
                        table2.Cell(Row=12, Column=2).Range.Text
    # 18 建筑高度
    result_list[18] = "" 
    # 19 更新时间
    result_list[19] = ""
    # 20备注
    result_list[20] = ""

    result = [s.replace('\r', '').replace('\x07', '') for s in result_list]
    doc.Close(SaveChanges=False)
    word.Quit()
    return result
def get_buildings_high(tech_check : Path):
    building_high = None
    wb = xlrd.open_workbook(str(tech_check))
    if "验收" in wb.sheet_names():
        ws = wb.sheet_by_name("验收")
        if "建筑高度" in str(ws.cell_value(26, 1)):
            building_high = str(ws.cell_value(26, 5))
        elif "建筑高度" in str(ws.cell_value(26, 2)):
            building_high = str(ws.cell_value(26, 6))
        elif "建筑高度" in str(ws.cell_value(27, 1)):
            building_high = str(ws.cell_value(27, 5))
        elif "建筑高度" in str(ws.cell_value(27, 2)):
            building_high = str(ws.cell_value(27, 6))
    return building_high

def validate_project(path_xls, fang_xls_path, fu_xls_path, filtered_name, progress_callback=None):
    filtered_filename = EXP_BASE_DIR.joinpath(filtered_name)

    wb_validate = xlrd.open_workbook(path_xls)
    ws_validate = wb_validate.sheet_by_index(0)
    project_list_total = [str(ws_validate.cell_value(i, 2)) for i in range(1, ws_validate.nrows)]
    wb_fang = xlrd.open_workbook(fang_xls_path)
    ws_fang = wb_fang.sheet_by_index(0)
    project_list_fang = [str(ws_fang.cell_value(i, 0)) for i in range(1, ws_fang.nrows)]
    wb_fu = xlrd.open_workbook(fu_xls_path)
    ws_fu = wb_fu.sheet_by_index(0)
    project_list_fu = [str(ws_fu.cell_value(i, 0)) for i in range(1, ws_fu.nrows)]
    project_fang_fu = project_list_fang + project_list_fu
    filtered = [item for item in project_list_total if item not in project_fang_fu]
    if (len(filtered) > 0) : 
        with open(filtered_filename, 'w+', encoding='utf-8') as f:
            for item in filtered:
                f.write(item+'\n')
    if progress_callback is not None:
        progress_callback(100, description="正在验证项目数量")
def main(fang_dir=None, fu_dir=None, validate_xls=None, progress_callback=None, args=None):
    if fang_dir is not None and fang_dir != "":
        make_fang_result(fang_dir=Path(fang_dir),progress_callback=progress_callback)
    if fu_dir is not None and fu_dir != "":
        make_fu_result(fu_dir=Path(fu_dir),progress_callback=progress_callback)
    if validate_xls is not None and validate_xls != "" :
        validate_project(path_xls=validate_xls,fu_xls_path="验收成果汇总表.xlsx", 
                         fang_xls_path="放线数据汇总.xlsx",
                        filtered_name="放线验收缺失的项目列表.txt",
                        progress_callback=progress_callback)

if __name__ == '__main__':
    # suply_make_fang(r"F:\专题库\原数据\放线txt补充")
    # pattern = r'\d{4}复\d{2}[A-Z]\d{3}'
    # path = r"F:\专题库\原数据\验收\2023复23B168\成果汇总表\成果汇总表_9.doc"
    # print(path)
    # match = re.search(pattern=pattern, string=path)
    # if match is not None:
    #         # 获取整个匹配的字符串
    #     full_match = match.group(0)
    #     print(match.string)
    #     print("Full match:", full_match)
    # else :
    #     print("No match found")
    wb_validate = load_workbook(r"验收成果汇总表.xlsx")
    ws_validate = wb_validate.worksheets[0]
    project_total = list(ws_validate.iter_rows(min_row=1, max_row=ws_validate.max_row, 
                                          min_col=1, max_col=1, values_only=True))
    print(type(project_total[0]), project_total[0])