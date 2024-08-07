import os 
import shutil
import glob
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook  
import xlrd3 as xlrd
import traceback
dir_fang = 'F:\专题库\原数据\\4.放线项目成果'
dir_fu = 'F:\专题库\原数据\\5.验收项目成果'

def check_dir(dir):
    dir = os.path.join(os.getcwd(), dir)
    if not os.path.exists(dir):
        os.makedirs(dir)
        return
    response = input(f"{dir} 已存在, 是否删除? (y/n): ").strip().lower()
    if response == 'y':
        shutil.rmtree(dir)
    os.makedirs(dir)
    return dir
def check_file(file, sheet_name="放线数据汇总"):
    file = os.path.join(os.getcwd(), file)
    if os.path.exists(file):
        os.remove(file)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(file)

def make_fang_result(excel_dir, sheet_name="放线数据汇总", save_name= "放线数据汇总.xlsx",
                title=["工程编号","建筑结构","建设单位","建设项目名称","建设位置",
                       "建设工程规划许可证号","更新时间","备注"],
                exp_dir="提取异常的放线excel",
                exception_filename='异常的文件列表.txt'):
    check_file(save_name, sheet_name=sheet_name)
    exp_dir = check_dir(exp_dir)    
    if os.path.exists(exception_filename):
        os.remove(exception_filename)

    excels = glob.glob(os.path.join(excel_dir, "*\\*放*.xls*"))
    excel_list = [s for s in excels if not os.path.basename(s).startswith('~')]
    wb = load_workbook(save_name)

    ws = wb[sheet_name]
    ws.append(title)
    ws.title= sheet_name
    exp_count = 1
    for excel_path in tqdm(excel_list, total=len(excel_list), unit="files", 
                        desc="正在写入excel"):
        try:
            result = get_fang_result(excel_path=excel_path)
        except Exception as e:
            with open(exception_filename, 'a', encoding='utf-8') as f:
                f.write(f"{exp_count}'\t'{excel_path}'\n'{traceback.format_exc()}")
                exp_count+=1
            
            copy_name = os.path.basename(os.path.dirname(excel_path)) + '-' + excel_path
            shutil.move(excel_path, os.path.join(excel_path, copy_name))
            continue
        ws.append(result)
    wb.save(save_name)

def get_fang_result(excel_path):
    result_list = [""] * 8
    wb = xlrd.open_workbook(excel_path)
    if "放线" in wb.sheet_names():
        sheet = wb.sheet_by_name('放线')
        result_list[0] = sheet.cell_value(1, 1)
        result_list[2] = sheet.cell_value(2, 1)
        result_list[3] = sheet.cell_value(4, 1)
        result_list[4] = sheet.cell_value(3, 1)
    return result_list
    
def temp():
    wb = xlrd.open_workbook(r'F:\专题库\原数据\\4.放线项目成果\\2023放23B161\\b161放线验收违法技术审查表.xlsx')
    sheet = wb.sheet_by_name('放线')
    result_list = [""] * 8
    if "放线" in wb.sheet_names():
        sheet = wb.sheet_by_name('放线')
        result_list[0] = sheet.cell_value(1, 1)
        result_list[2] = sheet.cell_value(2, 1)
        result_list[3] = sheet.cell_value(4, 1)
        result_list[4] = sheet.cell_value(3, 1)
    print(result_list)
if __name__ == '__main__':
    # copy_all_excel(dir_fang, "放线excel")
    # print(os.path.isdir('F:\\python_scripts\\放线excel'))
    # make_result()
    make_fang_result('F:\专题库\原数据\\4.放线项目成果')
    # temp()
    # get_excel_result("F:\\python_scripts\\放线excel\\2021放23B313-放线验收违法技术审查表.xls")